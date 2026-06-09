from __future__ import annotations

import cgi
import html
import json
import mimetypes
import sqlite3
import sys
import tempfile
import urllib.parse
from datetime import datetime
from html.parser import HTMLParser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import load_workbook


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
WEB_DIR = ROOT / "web"
DB_PATH = DATA_DIR / "webapp.sqlite3"


class HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] = []
        self._cell: list[str] = []
        self._in_td = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "tr":
            self._row = []
        if tag.lower() in {"td", "th"}:
            self._in_td = True
            self._cell = []
        if tag.lower() == "br" and self._in_td:
            self._cell.append(" ")

    def handle_data(self, data: str) -> None:
        if self._in_td:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"td", "th"}:
            value = "".join(self._cell).replace("\xa0", " ")
            self._row.append(" ".join(value.split()))
            self._in_td = False
        if tag.lower() == "tr" and self._row:
            self.rows.append(self._row)


def connect_db() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    with connect_db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS partners (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                phone TEXT NOT NULL DEFAULT '',
                contact_name TEXT NOT NULL DEFAULT '',
                memo TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS client_mappings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                hospital_name TEXT NOT NULL UNIQUE,
                partner_id INTEGER NOT NULL,
                memo TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                FOREIGN KEY (partner_id) REFERENCES partners(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS settlement_uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                settlement_month TEXT NOT NULL,
                filename TEXT NOT NULL,
                item_count INTEGER NOT NULL,
                hospital_count INTEGER NOT NULL,
                uploaded_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS prescription_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                settlement_month TEXT NOT NULL,
                hospital_name TEXT NOT NULL,
                product_name TEXT NOT NULL,
                maker_name TEXT NOT NULL DEFAULT '',
                insurance_code TEXT NOT NULL DEFAULT '',
                quantity TEXT NOT NULL DEFAULT '',
                prescription_amount TEXT NOT NULL DEFAULT '',
                uploaded_at TEXT NOT NULL,
                UNIQUE(settlement_month, hospital_name, product_name, maker_name, insurance_code)
            );
            """
        )


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def row_to_dict(row: sqlite3.Row) -> dict[str, object]:
    return {key: row[key] for key in row.keys()}


def decode_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "cp949", "euc-kr", "utf-16"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def find_header_row(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows):
        normalized = [clean_cell(value) for value in row]
        if "병의원명" in normalized and "제품명" in normalized:
            return idx, {name: pos for pos, name in enumerate(normalized)}
    raise ValueError("정산현황 파일에서 '병의원명'과 '제품명' 헤더를 찾지 못했습니다.")


def parse_settlement_rows(filename: str, data: bytes) -> tuple[str, list[dict[str, str]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx_settlement(data)
    return parse_html_settlement(data)


def parse_html_settlement(data: bytes) -> tuple[str, list[dict[str, str]]]:
    parser = HtmlTableParser()
    parser.feed(decode_bytes(data))
    header_idx, headers = find_header_row(parser.rows)
    return extract_items_from_rows(parser.rows[header_idx + 1 :], headers)


def parse_xlsx_settlement(data: bytes) -> tuple[str, list[dict[str, str]]]:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_file.write(data)
        temp_path = Path(temp_file.name)
    try:
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        ws = wb.active
        rows = [[clean_cell(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    finally:
        temp_path.unlink(missing_ok=True)
    header_idx, headers = find_header_row(rows)
    return extract_items_from_rows(rows[header_idx + 1 :], headers)


def extract_items_from_rows(rows: list[list[str]], headers: dict[str, int]) -> tuple[str, list[dict[str, str]]]:
    settlement_month = ""
    items: list[dict[str, str]] = []

    def value(row: list[str], header: str) -> str:
        idx = headers.get(header)
        if idx is None or idx >= len(row):
            return ""
        return clean_cell(row[idx])

    for row in rows:
        hospital = value(row, "병의원명")
        product = value(row, "제품명")
        month = value(row, "정산월")
        if month and not settlement_month:
            settlement_month = month
        if not hospital or not product:
            continue
        if "계" in hospital or product in {"", "-"}:
            continue
        items.append(
            {
                "settlement_month": month or settlement_month,
                "hospital_name": hospital,
                "product_name": product,
                "maker_name": value(row, "제약사명"),
                "insurance_code": value(row, "보험코드"),
                "quantity": value(row, "수량"),
                "prescription_amount": value(row, "처방금액"),
            }
        )

    settlement_month = settlement_month or datetime.now().strftime("%Y-%m")
    for item in items:
        item["settlement_month"] = item["settlement_month"] or settlement_month
    return settlement_month, items


def save_settlement(filename: str, items: list[dict[str, str]], settlement_month: str) -> dict[str, object]:
    uploaded_at = now_text()
    hospitals = {item["hospital_name"] for item in items}
    with connect_db() as conn:
        conn.execute("DELETE FROM prescription_items WHERE settlement_month = ?", (settlement_month,))
        conn.executemany(
            """
            INSERT OR IGNORE INTO prescription_items
                (settlement_month, hospital_name, product_name, maker_name, insurance_code,
                 quantity, prescription_amount, uploaded_at)
            VALUES
                (:settlement_month, :hospital_name, :product_name, :maker_name, :insurance_code,
                 :quantity, :prescription_amount, :uploaded_at)
            """,
            [{**item, "uploaded_at": uploaded_at} for item in items],
        )
        stored_item_count = conn.execute(
            "SELECT COUNT(*) FROM prescription_items WHERE settlement_month = ?",
            (settlement_month,),
        ).fetchone()[0]
        conn.execute(
            """
            INSERT INTO settlement_uploads
                (settlement_month, filename, item_count, hospital_count, uploaded_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (settlement_month, filename, stored_item_count, len(hospitals), uploaded_at),
        )
        conn.commit()
    return {
        "settlement_month": settlement_month,
        "item_count": stored_item_count,
        "hospital_count": len(hospitals),
        "uploaded_at": uploaded_at,
    }


def get_state() -> dict[str, object]:
    with connect_db() as conn:
        partners = [
            row_to_dict(row)
            for row in conn.execute("SELECT * FROM partners ORDER BY active DESC, name")
        ]
        clients = [
            row_to_dict(row)
            for row in conn.execute(
                """
                SELECT c.*, p.name AS partner_name, p.phone AS partner_phone
                FROM client_mappings c
                JOIN partners p ON p.id = c.partner_id
                ORDER BY c.active DESC, c.hospital_name
                """
            )
        ]
        uploads = [
            row_to_dict(row)
            for row in conn.execute(
                "SELECT * FROM settlement_uploads ORDER BY uploaded_at DESC LIMIT 10"
            )
        ]
        latest_month_row = conn.execute(
            "SELECT settlement_month FROM prescription_items ORDER BY uploaded_at DESC LIMIT 1"
        ).fetchone()
        latest_month = latest_month_row["settlement_month"] if latest_month_row else ""
        items = [
            row_to_dict(row)
            for row in conn.execute(
                """
                SELECT pi.*, cm.partner_id, p.name AS partner_name, p.phone AS partner_phone
                FROM prescription_items pi
                LEFT JOIN client_mappings cm ON cm.hospital_name = pi.hospital_name
                LEFT JOIN partners p ON p.id = cm.partner_id
                WHERE (? = '' OR pi.settlement_month = ?)
                ORDER BY pi.hospital_name, pi.product_name
                LIMIT 500
                """,
                (latest_month, latest_month),
            )
        ]
        unmapped = [
            row_to_dict(row)
            for row in conn.execute(
                """
                SELECT DISTINCT pi.hospital_name
                FROM prescription_items pi
                LEFT JOIN client_mappings cm ON cm.hospital_name = pi.hospital_name
                WHERE cm.id IS NULL
                ORDER BY pi.hospital_name
                """
            )
        ]
        stats = {
            "partner_count": conn.execute("SELECT COUNT(*) FROM partners").fetchone()[0],
            "client_count": conn.execute("SELECT COUNT(*) FROM client_mappings").fetchone()[0],
            "item_count": conn.execute("SELECT COUNT(*) FROM prescription_items").fetchone()[0],
            "unmapped_count": len(unmapped),
            "latest_month": latest_month,
        }
    return {
        "partners": partners,
        "clients": clients,
        "uploads": uploads,
        "items": items,
        "unmapped": unmapped,
        "stats": stats,
    }


def json_response(handler: BaseHTTPRequestHandler, payload: object, status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def read_json(handler: BaseHTTPRequestHandler) -> dict[str, object]:
    length = int(handler.headers.get("Content-Length", "0"))
    if length == 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw or "{}")


class WebAppHandler(BaseHTTPRequestHandler):
    def log_message(self, format: str, *args: object) -> None:
        print(f"{self.address_string()} - {format % args}")

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/state":
            json_response(self, get_state())
            return
        self.serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        try:
            if parsed.path == "/api/partners":
                self.create_partner()
            elif parsed.path == "/api/clients":
                self.create_client_mapping()
            elif parsed.path == "/api/settlements/upload":
                self.upload_settlement()
            else:
                json_response(self, {"error": "Not found"}, 404)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def do_DELETE(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed.query)
        try:
            if parsed.path == "/api/partners":
                target_id = int(params.get("id", ["0"])[0])
                with connect_db() as conn:
                    conn.execute("DELETE FROM partners WHERE id = ?", (target_id,))
                    conn.commit()
                json_response(self, {"ok": True})
            elif parsed.path == "/api/clients":
                target_id = int(params.get("id", ["0"])[0])
                with connect_db() as conn:
                    conn.execute("DELETE FROM client_mappings WHERE id = ?", (target_id,))
                    conn.commit()
                json_response(self, {"ok": True})
            else:
                json_response(self, {"error": "Not found"}, 404)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def serve_static(self, path: str) -> None:
        if path in {"", "/"}:
            path = "/index.html"
        safe_path = path.lstrip("/").replace("..", "")
        file_path = WEB_DIR / safe_path
        if not file_path.exists() or file_path.is_dir():
            self.send_error(404)
            return
        content = file_path.read_bytes()
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        if file_path.suffix == ".js":
            content_type = "application/javascript"
        if file_path.suffix == ".css":
            content_type = "text/css"
        self.send_response(200)
        self.send_header("Content-Type", f"{content_type}; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def create_partner(self) -> None:
        payload = read_json(self)
        name = clean_cell(payload.get("name"))
        if not name:
            raise ValueError("사업자명을 입력해 주세요.")
        with connect_db() as conn:
            conn.execute(
                """
                INSERT INTO partners (name, phone, contact_name, memo, active, created_at)
                VALUES (?, ?, ?, ?, 1, ?)
                ON CONFLICT(name) DO UPDATE SET
                    phone = excluded.phone,
                    contact_name = excluded.contact_name,
                    memo = excluded.memo,
                    active = 1
                """,
                (
                    name,
                    clean_cell(payload.get("phone")),
                    clean_cell(payload.get("contact_name")),
                    clean_cell(payload.get("memo")),
                    now_text(),
                ),
            )
            conn.commit()
        json_response(self, {"ok": True, "state": get_state()})

    def create_client_mapping(self) -> None:
        payload = read_json(self)
        hospital_name = clean_cell(payload.get("hospital_name"))
        partner_id = int(payload.get("partner_id") or 0)
        if not hospital_name:
            raise ValueError("병의원명을 입력해 주세요.")
        if not partner_id:
            raise ValueError("사업자를 선택해 주세요.")
        with connect_db() as conn:
            conn.execute(
                """
                INSERT INTO client_mappings (hospital_name, partner_id, memo, active, created_at)
                VALUES (?, ?, ?, 1, ?)
                ON CONFLICT(hospital_name) DO UPDATE SET
                    partner_id = excluded.partner_id,
                    memo = excluded.memo,
                    active = 1
                """,
                (hospital_name, partner_id, clean_cell(payload.get("memo")), now_text()),
            )
            conn.commit()
        json_response(self, {"ok": True, "state": get_state()})

    def upload_settlement(self) -> None:
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"})
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "file", None):
            raise ValueError("정산현황 파일을 선택해 주세요.")
        filename = Path(file_item.filename or "settlement.xls").name
        data = file_item.file.read()
        settlement_month, items = parse_settlement_rows(filename, data)
        if not items:
            raise ValueError("정산현황에서 병의원/제품 데이터를 찾지 못했습니다.")
        result = save_settlement(filename, items, settlement_month)
        json_response(self, {"ok": True, "result": result, "state": get_state()})


def main() -> int:
    init_db()
    host = "127.0.0.1"
    port = 8765
    server = ThreadingHTTPServer((host, port), WebAppHandler)
    print(f"Out-of-stock manager running at http://{host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("Server stopped")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
