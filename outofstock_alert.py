from __future__ import annotations

import argparse
import hashlib
import os
import re
import sqlite3
import sys
import unicodedata
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
DOWNLOAD_PATH = DATA_DIR / "outofstock-latest.pdf"
SALES_DOWNLOAD_PATH = DATA_DIR / "sales-list.xlsx"
DB_PATH = DATA_DIR / "alerts.sqlite3"
SALES_PATH = ROOT / "sales-list.xlsx"
DRIVE_FILE_ID = "15dOI-2gYbOLEett8Jfu4OWilAytZdM26"
DRIVE_DOWNLOAD_URL = (
    f"https://drive.google.com/uc?export=download&id={DRIVE_FILE_ID}"
)


@dataclass(frozen=True)
class SalesItem:
    client: str
    region: str
    product: str
    maker: str


@dataclass(frozen=True)
class Match:
    product: str
    clients: tuple[str, ...]
    maker: str
    matched_line: str
    match_type: str


def load_env(path: Path = ROOT / ".env") -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def google_drive_download_url(file_id: str) -> str:
    return f"https://drive.google.com/uc?export=download&id={file_id}"


def google_sheets_export_url(file_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"


def normalized_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "").upper()
    return re.sub(r"[^0-9A-Z가-힣]", "", text)


def product_stem(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "").upper()
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\[[^\]]*\]", "", text)
    text = re.sub(
        r"\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)*\s*(?:MG|ML|G|MCG|UG|IU|%|정|캡슐|T|C|B|관|병|포)?",
        "",
        text,
    )
    text = re.sub(r"(?:PTP|일반|다회용|일회용|신형|구형|서방|장용)", "", text)
    return re.sub(r"[^A-Z가-힣]", "", text)


def download_url(url: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "outofstock-alert/1.0"})

    with urllib.request.urlopen(req, timeout=60) as response:
        data = response.read()

    output_path.write_bytes(data)
    return output_path


def download_drive_pdf(url: str = DRIVE_DOWNLOAD_URL, output_path: Path = DOWNLOAD_PATH) -> Path:
    output_path = download_url(url, output_path)
    data = output_path.read_bytes()
    if not data.startswith(b"%PDF"):
        preview = data[:200].decode("utf-8", errors="replace")
        raise RuntimeError(
            "Google Drive에서 PDF가 아닌 응답을 받았습니다. "
            f"파일 공유 권한을 확인해 주세요. 응답 앞부분: {preview!r}"
        )

    return output_path


def resolve_sales_path(env: dict[str, str]) -> Path:
    if SALES_PATH.exists():
        return SALES_PATH

    sales_url = env.get("SALES_LIST_URL", "").strip()
    sales_file_id = env.get("SALES_LIST_FILE_ID", "")

    if sales_url:
        path = download_url(sales_url, SALES_DOWNLOAD_PATH)
        data = path.read_bytes()[:4]
        if not data.startswith(b"PK"):
            raise RuntimeError(
                "sales-list.xlsx를 다운로드했지만 엑셀 파일이 아닙니다. "
                "Google Drive 공유 권한 또는 SALES_LIST_FILE_ID를 확인해 주세요."
            )
        return path

    if sales_file_id:
        errors: list[str] = []
        for url in (
            google_drive_download_url(sales_file_id),
            google_sheets_export_url(sales_file_id),
        ):
            path = download_url(url, SALES_DOWNLOAD_PATH)
            if path.read_bytes()[:4].startswith(b"PK"):
                return path
            preview = path.read_bytes()[:120].decode("utf-8", errors="replace")
            errors.append(f"{url} -> {preview!r}")
        raise RuntimeError(
            "SALES_LIST_FILE_ID로 sales-list.xlsx를 다운로드하지 못했습니다. "
            "공유 권한을 '링크가 있는 모든 사용자 보기 가능'으로 설정했는지 확인해 주세요. "
            f"시도 결과: {' / '.join(errors)}"
        )

    raise RuntimeError(
        "sales-list.xlsx를 찾을 수 없습니다. "
        "로컬 실행이면 프로젝트 폴더에 sales-list.xlsx를 두고, "
        "GitHub Actions 실행이면 SALES_LIST_FILE_ID 또는 SALES_LIST_URL을 설정해 주세요."
    )


def extract_pdf_lines(pdf_path: Path) -> list[str]:
    reader = PdfReader(str(pdf_path))
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            clean = " ".join(line.strip().split())
            if clean:
                lines.append(clean)
    return lines


PROMOTION_STOP_RE = re.compile(
    r"(기본|기존|추가|프로모션|기간|대상|수수료|요율|전략|지급|신규|매출|처방시)"
)


def stockout_candidate_lines(pdf_path: Path) -> list[str]:
    reader = PdfReader(str(pdf_path))
    candidates: list[str] = []
    for page in reader.pages:
        page_started = False
        for raw_line in (page.extract_text() or "").splitlines():
            line = " ".join(raw_line.strip().split())
            if not line:
                continue
            if "제약사명" in line and "제품명" in line:
                page_started = True
                continue
            if not page_started:
                continue
            if PROMOTION_STOP_RE.search(line):
                break
            if line in {"출하 예정일", "내용"}:
                continue
            candidates.append(line)
    return candidates


def load_sales_items(path: Path = SALES_PATH) -> list[SalesItem]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}

    required = ["거래처명", "지역", "품목", "제약사"]
    missing = [name for name in required if name not in header_index]
    if missing:
        raise RuntimeError(f"sales-list.xlsx에 필요한 컬럼이 없습니다: {', '.join(missing)}")

    items: list[SalesItem] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        client = row[header_index["거래처명"]]
        product = row[header_index["품목"]]
        if not client or not product:
            continue
        items.append(
            SalesItem(
                client=str(client).strip(),
                region=str(row[header_index["지역"]] or "").strip(),
                product=str(product).strip(),
                maker=str(row[header_index["제약사"]] or "").strip(),
            )
        )
    return items


def find_matches(sales_items: list[SalesItem], pdf_lines: list[str]) -> list[Match]:
    by_product: dict[str, dict[str, object]] = {}
    for item in sales_items:
        entry = by_product.setdefault(
            item.product,
            {"clients": set(), "maker": item.maker},
        )
        clients = entry["clients"]
        assert isinstance(clients, set)
        clients.add(item.client)

    line_index = [
        (line, normalized_text(line), product_stem(line)) for line in pdf_lines
    ]

    matches: list[Match] = []
    for product, entry in by_product.items():
        full = normalized_text(product)
        stem = product_stem(product)
        if len(full) < 4 and len(stem) < 4:
            continue

        found_line = ""
        match_type = ""
        for line, line_full, line_stem in line_index:
            if full and len(full) >= 4 and full in line_full:
                found_line = line
                match_type = "정확/포함"
                break
            if stem and len(stem) >= 4 and stem in line_stem:
                found_line = line
                match_type = "제품명 기준"
                break

        if not found_line:
            continue

        clients = entry["clients"]
        assert isinstance(clients, set)
        matches.append(
            Match(
                product=product,
                clients=tuple(sorted(str(client) for client in clients)),
                maker=str(entry["maker"] or ""),
                matched_line=found_line,
                match_type=match_type,
            )
        )

    return sorted(matches, key=lambda item: (item.product, item.clients))


def connect_db(path: Path = DB_PATH) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sent_alerts (
            alert_hash TEXT PRIMARY KEY,
            sent_at TEXT NOT NULL,
            product TEXT NOT NULL,
            clients TEXT NOT NULL,
            matched_line TEXT NOT NULL
        )
        """
    )
    return conn


def alert_hash(match: Match) -> str:
    payload = "|".join([match.product, ",".join(match.clients), match.matched_line])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def no_match_alert_hash(day: str | None = None) -> str:
    target_day = day or datetime.now().strftime("%Y-%m-%d")
    return hashlib.sha256(f"NO_MATCH|{target_day}".encode("utf-8")).hexdigest()


def has_sent_hash(conn: sqlite3.Connection, hash_value: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sent_alerts WHERE alert_hash = ?",
        (hash_value,),
    ).fetchone()
    return row is not None


def unsent_matches(matches: list[Match], conn: sqlite3.Connection) -> list[Match]:
    result: list[Match] = []
    for match in matches:
        if not has_sent_hash(conn, alert_hash(match)):
            result.append(match)
    return result


def mark_sent(matches: list[Match], conn: sqlite3.Connection) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    conn.executemany(
        """
        INSERT OR IGNORE INTO sent_alerts
            (alert_hash, sent_at, product, clients, matched_line)
        VALUES (?, ?, ?, ?, ?)
        """,
        [
            (
                alert_hash(match),
                now,
                match.product,
                ", ".join(match.clients),
                match.matched_line,
            )
            for match in matches
        ],
    )
    conn.commit()


def mark_no_match_sent(conn: sqlite3.Connection) -> None:
    today = datetime.now().strftime("%Y-%m-%d")
    conn.execute(
        """
        INSERT OR IGNORE INTO sent_alerts
            (alert_hash, sent_at, product, clients, matched_line)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            no_match_alert_hash(today),
            datetime.now().isoformat(timespec="seconds"),
            "__NO_MATCH__",
            "",
            today,
        ),
    )
    conn.commit()


def format_summary(matches: list[Match], limit: int = 20) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    if not matches:
        return f"[품절 알림]\n확인일: {today}\n현재 거래처 품목과 매칭된 품절 품목이 없습니다."

    lines = [
        "[품절 알림]",
        f"확인일: {today}",
        f"매칭 품목: {len(matches)}개",
        "",
    ]
    for idx, match in enumerate(matches[:limit], start=1):
        clients = ", ".join(match.clients[:5])
        if len(match.clients) > 5:
            clients += f" 외 {len(match.clients) - 5}곳"
        lines.extend(
            [
                f"{idx}. {match.product}",
                f"   거래처: {clients}",
                f"   PDF 항목: {match.matched_line}",
                f"   매칭: {match.match_type}",
            ]
        )

    if len(matches) > limit:
        lines.append(f"외 {len(matches) - limit}개 품목이 더 있습니다.")
    return "\n".join(lines)


def split_message(text: str, limit: int = 3900) -> list[str]:
    chunks: list[str] = []
    current: list[str] = []
    current_len = 0
    for line in text.splitlines():
        next_len = current_len + len(line) + 1
        if current and next_len > limit:
            chunks.append("\n".join(current))
            current = [line]
            current_len = len(line) + 1
        else:
            current.append(line)
            current_len = next_len
    if current:
        chunks.append("\n".join(current))
    return chunks


def send_telegram_message(token: str, chat_id: str, text: str) -> None:
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = urllib.parse.urlencode({"chat_id": chat_id, "text": text}).encode("utf-8")
    req = urllib.request.Request(url, data=payload, method="POST")

    with urllib.request.urlopen(req, timeout=30) as response:
        body = response.read().decode("utf-8", errors="replace")

    if '"ok":true' not in body:
        raise RuntimeError(f"텔레그램 발송 실패: {body}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="품절 리스트 매칭 및 텔레그램 알림")
    parser.add_argument("--dry-run", action="store_true", help="발송하지 않고 결과만 출력")
    parser.add_argument("--send-test", action="store_true", help="텔레그램 테스트 메시지만 발송")
    parser.add_argument("--send", action="store_true", help="매칭된 신규 품절 알림 발송")
    parser.add_argument(
        "--use-existing-pdf",
        action="store_true",
        help="Drive 다운로드 대신 data/outofstock-latest.pdf 사용",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    env = {**os.environ, **load_env()}
    token = env.get("TELEGRAM_BOT_TOKEN", "")
    chat_id = env.get("TELEGRAM_CHAT_ID", "")

    if args.send_test:
        if not token or not chat_id:
            raise RuntimeError(".env에 TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID가 필요합니다.")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        send_telegram_message(
            token,
            chat_id,
            f"[품절 알림봇 테스트]\n텔레그램 연결이 정상입니다.\n발송시각: {now}",
        )
        print("telegram_test_sent=true")
        return 0

    sales_path = resolve_sales_path(env)

    pdf_path = DOWNLOAD_PATH if args.use_existing_pdf else download_drive_pdf()
    all_pdf_lines = extract_pdf_lines(pdf_path)
    pdf_lines = stockout_candidate_lines(pdf_path)
    sales_items = load_sales_items(sales_path)
    matches = find_matches(sales_items, pdf_lines)

    print(f"pdf_path={pdf_path}")
    print(f"sales_path={sales_path}")
    print(f"pdf_lines={len(all_pdf_lines)}")
    print(f"stockout_candidate_lines={len(pdf_lines)}")
    print(f"sales_items={len(sales_items)}")
    print(f"matches={len(matches)}")
    print(format_summary(matches, limit=10))

    if args.dry_run:
        return 0

    if not args.send:
        print("발송하려면 --send 옵션을 사용하세요.")
        return 0

    if not token or not chat_id:
        raise RuntimeError(".env에 TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID가 필요합니다.")

    with connect_db() as conn:
        if not matches:
            no_match_hash = no_match_alert_hash()
            if has_sent_hash(conn, no_match_hash):
                print("no_match_alert_already_sent_today=true")
                return 0
            send_telegram_message(token, chat_id, format_summary(matches))
            mark_no_match_sent(conn)
            print("no_match_alert_sent=true")
            return 0

        new_matches = unsent_matches(matches, conn)
        if not new_matches:
            print("new_matches=0")
            return 0
        for message in split_message(format_summary(new_matches)):
            send_telegram_message(token, chat_id, message)
        mark_sent(new_matches, conn)
        print(f"sent_matches={len(new_matches)}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
